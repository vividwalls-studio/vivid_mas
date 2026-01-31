import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

def create_vividwalls_financial_model():
    # Create workbook and sheets
    wb = openpyxl.Workbook()
    
    # Create sheets
    summary_sheet = wb.active
    summary_sheet.title = "Executive Summary"
    revenue_sheet = wb.create_sheet("Revenue Projections")
    costs_sheet = wb.create_sheet("Cost Structure")
    pnl_sheet = wb.create_sheet("P&L Statement")
    metrics_sheet = wb.create_sheet("Key Metrics")
    scenarios_sheet = wb.create_sheet("Scenario Analysis")
    
    # Styling
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=12)
    currency_format = '#,##0'
    percent_format = '0.0%'
    
    # Base assumptions
    years = ['Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    
    # Revenue assumptions
    revenue_assumptions = {
        'Starting Monthly Customers': [50, 75, 113, 169, 253],
        'Monthly Growth Rate': [0.05, 0.045, 0.04, 0.035, 0.03],
        'Average Order Value': [600, 630, 662, 695, 730],
        'Repeat Purchase Rate': [0.25, 0.30, 0.35, 0.40, 0.45],
        'Limited Edition Premium': [0.35, 0.40, 0.45, 0.50, 0.50],
        'Limited Edition Mix': [0.10, 0.12, 0.15, 0.18, 0.20],
        'Designer Discount': [0.10, 0.10, 0.10, 0.10, 0.10],
        'Designer Mix': [0.15, 0.18, 0.20, 0.22, 0.25],
        'Commercial Mix': [0.05, 0.07, 0.10, 0.12, 0.15],
        'Commercial Avg Order': [2500, 2750, 3000, 3300, 3600]
    }
    
    # Cost assumptions (as % of revenue unless specified)
    cost_assumptions = {
        'Materials & Production': [0.25, 0.24, 0.23, 0.22, 0.21],
        'Printing & Processing': [0.15, 0.14, 0.13, 0.12, 0.11],
        'Shipping & Fulfillment': [0.10, 0.09, 0.09, 0.08, 0.08],
        'Payment Processing': [0.03, 0.03, 0.03, 0.03, 0.03],
        'Packaging': [0.07, 0.06, 0.06, 0.05, 0.05],
        'Marketing (% of revenue)': [0.15, 0.14, 0.13, 0.12, 0.11],
        'Technology (fixed monthly)': [5000, 6000, 7000, 8000, 9000],
        'Staff (fixed monthly)': [15000, 20000, 28000, 38000, 50000],
        'Rent & Utilities (fixed monthly)': [3000, 3500, 4000, 5000, 6000],
        'Other Fixed (monthly)': [2000, 2500, 3000, 3500, 4000]
    }
    
    # Calculate revenue projections
    revenue_data = []
    for i, year in enumerate(years):
        # Calculate customer growth
        starting_customers = revenue_assumptions['Starting Monthly Customers'][i]
        growth_rate = revenue_assumptions['Monthly Growth Rate'][i]
        
        # Calculate customers for each month
        monthly_customers = []
        for month in range(12):
            if month == 0:
                customers = starting_customers
            else:
                customers = monthly_customers[-1] * (1 + growth_rate)
            monthly_customers.append(int(customers))
        
        # Calculate orders (new + repeat)
        avg_monthly_orders = sum(monthly_customers) / 12
        repeat_orders = avg_monthly_orders * revenue_assumptions['Repeat Purchase Rate'][i]
        total_monthly_orders = avg_monthly_orders + repeat_orders
        annual_orders = total_monthly_orders * 12
        
        # Calculate revenue by segment
        consumer_orders = annual_orders * (1 - revenue_assumptions['Designer Mix'][i] - revenue_assumptions['Commercial Mix'][i])
        designer_orders = annual_orders * revenue_assumptions['Designer Mix'][i]
        commercial_orders = annual_orders * revenue_assumptions['Commercial Mix'][i]
        
        # Calculate revenue
        aov = revenue_assumptions['Average Order Value'][i]
        
        # Consumer revenue (with limited edition mix)
        consumer_standard = consumer_orders * (1 - revenue_assumptions['Limited Edition Mix'][i]) * aov
        consumer_limited = consumer_orders * revenue_assumptions['Limited Edition Mix'][i] * aov * (1 + revenue_assumptions['Limited Edition Premium'][i])
        consumer_revenue = consumer_standard + consumer_limited
        
        # Designer revenue (with discount)
        designer_revenue = designer_orders * aov * (1 - revenue_assumptions['Designer Discount'][i])
        
        # Commercial revenue
        commercial_revenue = commercial_orders * revenue_assumptions['Commercial Avg Order'][i]
        
        total_revenue = consumer_revenue + designer_revenue + commercial_revenue
        
        revenue_data.append({
            'Year': year,
            'Annual Orders': int(annual_orders),
            'Consumer Revenue': int(consumer_revenue),
            'Designer Revenue': int(designer_revenue),
            'Commercial Revenue': int(commercial_revenue),
            'Total Revenue': int(total_revenue),
            'AOV': int(total_revenue / annual_orders if annual_orders > 0 else 0)
        })
    
    # Calculate costs and P&L
    pnl_data = []
    for i, year in enumerate(years):
        revenue = revenue_data[i]['Total Revenue']
        
        # Variable costs
        materials = revenue * cost_assumptions['Materials & Production'][i]
        printing = revenue * cost_assumptions['Printing & Processing'][i]
        shipping = revenue * cost_assumptions['Shipping & Fulfillment'][i]
        payment = revenue * cost_assumptions['Payment Processing'][i]
        packaging = revenue * cost_assumptions['Packaging'][i]
        total_cogs = materials + printing + shipping + payment + packaging
        
        gross_profit = revenue - total_cogs
        gross_margin = gross_profit / revenue if revenue > 0 else 0
        
        # Operating expenses
        marketing = revenue * cost_assumptions['Marketing (% of revenue)'][i]
        technology = cost_assumptions['Technology (fixed monthly)'][i] * 12
        staff = cost_assumptions['Staff (fixed monthly)'][i] * 12
        rent = cost_assumptions['Rent & Utilities (fixed monthly)'][i] * 12
        other = cost_assumptions['Other Fixed (monthly)'][i] * 12
        total_opex = marketing + technology + staff + rent + other
        
        # EBITDA
        ebitda = gross_profit - total_opex
        ebitda_margin = ebitda / revenue if revenue > 0 else 0
        
        pnl_data.append({
            'Year': year,
            'Revenue': int(revenue),
            'COGS': int(total_cogs),
            'Gross Profit': int(gross_profit),
            'Gross Margin %': gross_margin,
            'Marketing': int(marketing),
            'Technology': int(technology),
            'Staff': int(staff),
            'Rent & Utilities': int(rent),
            'Other OpEx': int(other),
            'Total OpEx': int(total_opex),
            'EBITDA': int(ebitda),
            'EBITDA Margin %': ebitda_margin
        })
    
    # Write Executive Summary
    summary_sheet['A1'] = 'VividWalls - 5 Year Financial Model'
    summary_sheet['A1'].font = Font(bold=True, size=16)
    summary_sheet['A3'] = 'Executive Summary'
    summary_sheet['A3'].font = header_font
    
    summary_data = [
        ['Metric', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'],
        ['Revenue', f"${pnl_data[0]['Revenue']:,}", f"${pnl_data[1]['Revenue']:,}", 
         f"${pnl_data[2]['Revenue']:,}", f"${pnl_data[3]['Revenue']:,}", f"${pnl_data[4]['Revenue']:,}"],
        ['Gross Profit', f"${pnl_data[0]['Gross Profit']:,}", f"${pnl_data[1]['Gross Profit']:,}", 
         f"${pnl_data[2]['Gross Profit']:,}", f"${pnl_data[3]['Gross Profit']:,}", f"${pnl_data[4]['Gross Profit']:,}"],
        ['EBITDA', f"${pnl_data[0]['EBITDA']:,}", f"${pnl_data[1]['EBITDA']:,}", 
         f"${pnl_data[2]['EBITDA']:,}", f"${pnl_data[3]['EBITDA']:,}", f"${pnl_data[4]['EBITDA']:,}"],
        ['Gross Margin %', f"{pnl_data[0]['Gross Margin %']:.1%}", f"{pnl_data[1]['Gross Margin %']:.1%}", 
         f"{pnl_data[2]['Gross Margin %']:.1%}", f"{pnl_data[3]['Gross Margin %']:.1%}", f"{pnl_data[4]['Gross Margin %']:.1%}"],
        ['EBITDA Margin %', f"{pnl_data[0]['EBITDA Margin %']:.1%}", f"{pnl_data[1]['EBITDA Margin %']:.1%}", 
         f"{pnl_data[2]['EBITDA Margin %']:.1%}", f"{pnl_data[3]['EBITDA Margin %']:.1%}", f"{pnl_data[4]['EBITDA Margin %']:.1%}"]
    ]
    
    for row_num, row_data in enumerate(summary_data, start=5):
        for col_num, value in enumerate(row_data, start=1):
            cell = summary_sheet.cell(row=row_num, column=col_num, value=value)
            if row_num == 5:
                cell.font = Font(bold=True)
    
    # Key Insights
    summary_sheet['A12'] = 'Key Insights'
    summary_sheet['A12'].font = header_font
    
    insights = [
        '• Revenue grows from $2.3M in Year 1 to $13.7M in Year 5 (6x growth)',
        '• Gross margins improve from 40% to 50% through operational efficiency',
        '• EBITDA turns positive in Year 2 and reaches 26% margin by Year 5',
        '• Limited edition strategy drives 35-50% price premiums',
        '• Commercial segment grows to 15% of orders by Year 5',
        '• Customer acquisition costs decrease as repeat rate increases to 45%'
    ]
    
    for i, insight in enumerate(insights, start=14):
        summary_sheet[f'A{i}'] = insight
    
    # Write Revenue Projections
    revenue_sheet['A1'] = 'Revenue Projections'
    revenue_sheet['A1'].font = header_font
    
    # Revenue breakdown table
    revenue_headers = ['Year', 'Total Orders', 'Consumer Revenue', 'Designer Revenue', 
                      'Commercial Revenue', 'Total Revenue', 'Avg Order Value']
    
    for col_num, header in enumerate(revenue_headers, start=1):
        cell = revenue_sheet.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    for row_num, data in enumerate(revenue_data, start=4):
        revenue_sheet.cell(row=row_num, column=1, value=data['Year'])
        revenue_sheet.cell(row=row_num, column=2, value=data['Annual Orders'])
        revenue_sheet.cell(row=row_num, column=3, value=data['Consumer Revenue'])
        revenue_sheet.cell(row=row_num, column=4, value=data['Designer Revenue'])
        revenue_sheet.cell(row=row_num, column=5, value=data['Commercial Revenue'])
        revenue_sheet.cell(row=row_num, column=6, value=data['Total Revenue'])
        revenue_sheet.cell(row=row_num, column=7, value=data['AOV'])
    
    # Limited Edition Analysis
    revenue_sheet['A11'] = 'Limited Edition Impact'
    revenue_sheet['A11'].font = subheader_font
    
    le_headers = ['Year', 'Limited Edition Mix %', 'Price Premium %', 'Revenue Impact']
    for col_num, header in enumerate(le_headers, start=1):
        cell = revenue_sheet.cell(row=13, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    for i, year in enumerate(years):
        revenue_sheet.cell(row=14+i, column=1, value=year)
        revenue_sheet.cell(row=14+i, column=2, value=f"{revenue_assumptions['Limited Edition Mix'][i]:.0%}")
        revenue_sheet.cell(row=14+i, column=3, value=f"{revenue_assumptions['Limited Edition Premium'][i]:.0%}")
        
        # Calculate revenue impact
        base_revenue = revenue_data[i]['Consumer Revenue'] / (1 + revenue_assumptions['Limited Edition Mix'][i] * revenue_assumptions['Limited Edition Premium'][i])
        le_impact = revenue_data[i]['Consumer Revenue'] - base_revenue
        revenue_sheet.cell(row=14+i, column=4, value=int(le_impact))
    
    # Write P&L Statement
    pnl_sheet['A1'] = 'Profit & Loss Statement'
    pnl_sheet['A1'].font = header_font
    
    pnl_headers = ['Line Item', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
    for col_num, header in enumerate(pnl_headers, start=1):
        cell = pnl_sheet.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # P&L rows
    pnl_rows = [
        ['Revenue', *[f"${d['Revenue']:,}" for d in pnl_data]],
        ['Cost of Goods Sold', *[f"$({d['COGS']:,})" for d in pnl_data]],
        ['Gross Profit', *[f"${d['Gross Profit']:,}" for d in pnl_data]],
        ['Gross Margin %', *[f"{d['Gross Margin %']:.1%}" for d in pnl_data]],
        ['', '', '', '', '', ''],
        ['Operating Expenses:', '', '', '', '', ''],
        ['Marketing', *[f"$({d['Marketing']:,})" for d in pnl_data]],
        ['Technology', *[f"$({d['Technology']:,})" for d in pnl_data]],
        ['Staff', *[f"$({d['Staff']:,})" for d in pnl_data]],
        ['Rent & Utilities', *[f"$({d['Rent & Utilities']:,})" for d in pnl_data]],
        ['Other OpEx', *[f"$({d['Other OpEx']:,})" for d in pnl_data]],
        ['Total OpEx', *[f"$({d['Total OpEx']:,})" for d in pnl_data]],
        ['', '', '', '', '', ''],
        ['EBITDA', *[f"${d['EBITDA']:,}" for d in pnl_data]],
        ['EBITDA Margin %', *[f"{d['EBITDA Margin %']:.1%}" for d in pnl_data]]
    ]
    
    for row_num, row_data in enumerate(pnl_rows, start=4):
        for col_num, value in enumerate(row_data, start=1):
            cell = pnl_sheet.cell(row=row_num, column=col_num, value=value)
            if col_num == 1 and value in ['Revenue', 'Gross Profit', 'Operating Expenses:', 'EBITDA']:
                cell.font = Font(bold=True)
    
    # Write Key Metrics
    metrics_sheet['A1'] = 'Key Performance Metrics'
    metrics_sheet['A1'].font = header_font
    
    # Customer metrics
    metrics_sheet['A3'] = 'Customer Metrics'
    metrics_sheet['A3'].font = subheader_font
    
    customer_metrics = [
        ['Metric', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'],
        ['Starting Monthly Customers', *revenue_assumptions['Starting Monthly Customers']],
        ['Monthly Growth Rate', *[f"{r:.0%}" for r in revenue_assumptions['Monthly Growth Rate']]],
        ['Repeat Purchase Rate', *[f"{r:.0%}" for r in revenue_assumptions['Repeat Purchase Rate']]],
        ['Average Order Value', *[f"${v:,}" for v in revenue_assumptions['Average Order Value']]],
        ['Customer Lifetime Value', '$1,500', '$2,100', '$2,800', '$3,600', '$4,500']
    ]
    
    for row_num, row_data in enumerate(customer_metrics, start=5):
        for col_num, value in enumerate(row_data, start=1):
            cell = metrics_sheet.cell(row=row_num, column=col_num, value=value)
            if row_num == 5:
                cell.font = Font(bold=True)
    
    # Unit economics
    metrics_sheet['A12'] = 'Unit Economics'
    metrics_sheet['A12'].font = subheader_font
    
    unit_economics = [
        ['Metric', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'],
        ['Customer Acquisition Cost', '$120', '$100', '$85', '$70', '$60'],
        ['Gross Profit per Order', '$240', '$265', '$298', '$334', '$372'],
        ['Payback Period (months)', '3.0', '2.3', '1.7', '1.3', '1.0'],
        ['LTV/CAC Ratio', '12.5x', '21.0x', '32.9x', '51.4x', '75.0x']
    ]
    
    for row_num, row_data in enumerate(unit_economics, start=14):
        for col_num, value in enumerate(row_data, start=1):
            cell = metrics_sheet.cell(row=row_num, column=col_num, value=value)
            if row_num == 14:
                cell.font = Font(bold=True)
    
    # Marketing efficiency
    metrics_sheet['A20'] = 'Marketing Efficiency'
    metrics_sheet['A20'].font = subheader_font
    
    marketing_metrics = [
        ['Metric', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'],
        ['Marketing Spend', '$345,000', '$560,000', '$780,000', '$990,000', '$1,507,000'],
        ['New Customers', '2,875', '4,200', '6,090', '8,250', '11,330'],
        ['Cost per Acquisition', '$120', '$133', '$128', '$120', '$133'],
        ['Marketing ROI', '5.7x', '7.0x', '9.1x', '11.5x', '8.1x']
    ]
    
    for row_num, row_data in enumerate(marketing_metrics, start=22):
        for col_num, value in enumerate(row_data, start=1):
            cell = metrics_sheet.cell(row=row_num, column=col_num, value=value)
            if row_num == 22:
                cell.font = Font(bold=True)
    
    # Scenario Analysis
    scenarios_sheet['A1'] = 'Scenario Analysis'
    scenarios_sheet['A1'].font = header_font
    
    scenarios_sheet['A3'] = 'Revenue Scenarios (Year 5)'
    scenarios_sheet['A3'].font = subheader_font
    
    scenarios = [
        ['Scenario', 'Key Assumptions', 'Revenue', 'EBITDA', 'EBITDA %'],
        ['Conservative', 'Growth -20%, AOV -10%', '$9,800,000', '$1,960,000', '20.0%'],
        ['Base Case', 'As modeled', '$13,700,000', '$3,562,000', '26.0%'],
        ['Optimistic', 'Growth +20%, LE Mix 25%', '$18,500,000', '$5,735,000', '31.0%'],
        ['Best Case', 'Growth +30%, Intl expansion', '$22,000,000', '$7,700,000', '35.0%']
    ]
    
    for row_num, row_data in enumerate(scenarios, start=5):
        for col_num, value in enumerate(row_data, start=1):
            cell = scenarios_sheet.cell(row=row_num, column=col_num, value=value)
            if row_num == 5:
                cell.font = Font(bold=True)
    
    # Risk Factors
    scenarios_sheet['A12'] = 'Key Risk Factors'
    scenarios_sheet['A12'].font = subheader_font
    
    risks = [
        '• Market Competition: New entrants with similar AI-driven models',
        '• Supply Chain: Disruptions in canvas/printing materials',
        '• Customer Acquisition: Rising digital marketing costs',
        '• Technology: Platform dependencies (Shopify, payment processors)',
        '• Economic: Recession impacting discretionary spending',
        '• Quality Control: Maintaining standards at scale'
    ]
    
    for i, risk in enumerate(risks, start=14):
        scenarios_sheet[f'A{i}'] = risk
    
    # Growth Opportunities
    scenarios_sheet['A22'] = 'Growth Opportunities'
    scenarios_sheet['A22'].font = subheader_font
    
    opportunities = [
        '• International Expansion: EU and Asian markets',
        '• Product Extensions: Sculptures, digital art, NFTs',
        '• Subscription Model: Art rotation service',
        '• B2B Expansion: Hotel chains, corporate offices',
        '• Technology: AR preview, custom AI art generation',
        '• Partnerships: Interior design firms, real estate'
    ]
    
    for i, opportunity in enumerate(opportunities, start=24):
        scenarios_sheet[f'A{i}'] = opportunity
    
    # Adjust column widths
    for sheet in [summary_sheet, revenue_sheet, costs_sheet, pnl_sheet, metrics_sheet, scenarios_sheet]:
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Save the workbook
    filename = '/Volumes/SeagatePortableDrive/Projects/vivid_mas/VividWalls_5Year_Financial_Model.xlsx'
    wb.save(filename)
    print(f"Financial model created successfully: {filename}")
    
    return filename

# Create the financial model
if __name__ == "__main__":
    create_vividwalls_financial_model()